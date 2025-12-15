import sys
import tempfile
import unittest
from contextlib import redirect_stdout
from io import StringIO
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "src"
sys.path.insert(0, str(SRC))

from clearml_dataset_excel.cli import main  # noqa: E402


class TestCliRunNoUploadStagePersists(unittest.TestCase):
    def test_run_no_upload_keeps_stage_dir(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            data = base / "data"
            data.mkdir()
            meas = data / "meas.csv"
            meas.write_text("time,value\n0,1\n1,2\n", encoding="utf-8")

            (base / "clearml_dataset_excel_runner.exe").write_text("dummy", encoding="utf-8")

            spec_path = base / "spec.yaml"
            spec_path.write_text(
                "\n".join(
                    [
                        "schema_version: 1",
                        "addin:",
                        "  enabled: true",
                        "condition:",
                        "  columns:",
                        "    - {name: id, type: str}",
                        "    - {name: meas_path, type: path}",
                        "files:",
                        "  - id: m",
                        "    path_column: meas_path",
                        "    mapping: {axes: {t: time}, targets: [{name: f, source: value}]}",
                    ]
                )
                + "\n",
                encoding="utf-8",
            )

            import openpyxl

            xlsm = base / "conditions.xlsm"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Conditions"
            ws.append(["id", "meas_path"])
            ws.append(["s1", "data/meas.csv"])
            wb.create_sheet("Extra")["A1"].value = "keep"
            wb.save(xlsm)

            buf = StringIO()
            with redirect_stdout(buf):
                rc = main(
                    [
                        "run",
                        "--spec",
                        spec_path.as_posix(),
                        "--excel",
                        xlsm.as_posix(),
                        "--output-root",
                        base.as_posix(),
                        "--no-upload",
                    ]
                )

            self.assertEqual(rc, 0)
            lines = [ln.strip() for ln in buf.getvalue().splitlines() if ln.strip()]
            self.assertGreaterEqual(len(lines), 2)
            output_dir = Path(lines[0]).expanduser().resolve()
            stage_dir = Path(lines[1]).expanduser().resolve()
            self.assertTrue(output_dir.exists())
            self.assertTrue(stage_dir.exists())
            self.assertTrue((stage_dir / "payload.json").exists())
            self.assertTrue((stage_dir / "template" / "clearml_dataset_excel_runner.exe").exists())
            self.assertTrue(str(stage_dir).startswith(str(output_dir)))

            import openpyxl

            tmpl = stage_dir / "template" / "condition_template.xlsm"
            self.assertTrue(tmpl.exists())
            wb2 = openpyxl.load_workbook(tmpl)
            self.assertIn("Extra", wb2.sheetnames)
            self.assertEqual(wb2["Conditions"].max_row, 1)

            import json

            payload = json.loads((stage_dir / "payload.json").read_text(encoding="utf-8"))
            self.assertEqual(payload.get("runner_exe_windows"), "template/clearml_dataset_excel_runner.exe")


if __name__ == "__main__":
    unittest.main()
