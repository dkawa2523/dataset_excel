import sys
import tempfile
import unittest
from pathlib import Path
from xml.etree import ElementTree as ET

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "src"
sys.path.insert(0, str(SRC))

from clearml_dataset_excel.vba_embedder import embed_vba_module_into_xlsm  # noqa: E402


def _add_vba_project_bin(xlsm_path: Path, data: bytes) -> None:
    import zipfile

    with zipfile.ZipFile(xlsm_path, "a") as z:
        z.writestr("xl/vbaProject.bin", data)


def _read_zip_entry(xlsm_path: Path, name: str) -> bytes:
    import zipfile

    with zipfile.ZipFile(xlsm_path, "r") as z:
        return z.read(name)


class TestVbaEmbedderTemplateExcel(unittest.TestCase):
    def test_embed_from_template_injects_vba_and_patches_xml(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            target = base / "target.xlsm"
            donor = base / "donor.xlsm"

            import openpyxl

            openpyxl.Workbook().save(target)
            openpyxl.Workbook().save(donor)

            donor_vba = b"dummy ClearMLDatasetExcel_Run dummy"
            _add_vba_project_bin(donor, donor_vba)

            embed_vba_module_into_xlsm(excel_path=target, template_excel=donor, overwrite=True)

            names = _read_zip_entry(target, "[Content_Types].xml")
            self.assertIn(b"vbaProject.bin", names)
            self.assertEqual(_read_zip_entry(target, "xl/vbaProject.bin"), donor_vba)

            # Content types: workbook is macro-enabled and vbaProject.bin is registered.
            ct_root = ET.fromstring(_read_zip_entry(target, "[Content_Types].xml"))
            ct_ns = {"ct": "http://schemas.openxmlformats.org/package/2006/content-types"}
            workbook_ct = None
            vba_ct = None
            for o in ct_root.findall("ct:Override", ct_ns):
                if o.attrib.get("PartName") == "/xl/workbook.xml":
                    workbook_ct = o.attrib.get("ContentType")
                if o.attrib.get("PartName") == "/xl/vbaProject.bin":
                    vba_ct = o.attrib.get("ContentType")
            self.assertEqual(workbook_ct, "application/vnd.ms-excel.sheet.macroEnabled.main+xml")
            self.assertEqual(vba_ct, "application/vnd.ms-office.vbaProject")

            # Relationship: workbook.xml.rels contains vbaProject relation.
            rel_root = ET.fromstring(_read_zip_entry(target, "xl/_rels/workbook.xml.rels"))
            rel_ns = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
            rel_types = [r.attrib.get("Type") for r in rel_root.findall("r:Relationship", rel_ns)]
            self.assertIn("http://schemas.microsoft.com/office/2006/relationships/vbaProject", rel_types)

            # workbook.xml: ensure codeName is set for robust ThisWorkbook usage.
            wb_root = ET.fromstring(_read_zip_entry(target, "xl/workbook.xml"))
            wb_ns = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
            wbpr = wb_root.find("s:workbookPr", wb_ns)
            self.assertIsNotNone(wbpr)
            assert wbpr is not None
            self.assertEqual(wbpr.attrib.get("codeName"), "ThisWorkbook")

    def test_embed_from_template_respects_overwrite_flag(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            target = base / "target.xlsm"
            donor1 = base / "donor1.xlsm"
            donor2 = base / "donor2.xlsm"

            import openpyxl

            openpyxl.Workbook().save(target)
            openpyxl.Workbook().save(donor1)
            openpyxl.Workbook().save(donor2)

            v1 = b"ONE ClearMLDatasetExcel_Run"
            v2 = b"TWO ClearMLDatasetExcel_Run"
            _add_vba_project_bin(donor1, v1)
            _add_vba_project_bin(donor2, v2)

            embed_vba_module_into_xlsm(excel_path=target, template_excel=donor1, overwrite=True)
            self.assertEqual(_read_zip_entry(target, "xl/vbaProject.bin"), v1)

            embed_vba_module_into_xlsm(excel_path=target, template_excel=donor2, overwrite=False)
            self.assertEqual(_read_zip_entry(target, "xl/vbaProject.bin"), v1)

            embed_vba_module_into_xlsm(excel_path=target, template_excel=donor2, overwrite=True)
            self.assertEqual(_read_zip_entry(target, "xl/vbaProject.bin"), v2)

    def test_embed_from_template_requires_expected_symbol(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            target = base / "target.xlsm"
            donor = base / "donor.xlsm"

            import openpyxl

            openpyxl.Workbook().save(target)
            openpyxl.Workbook().save(donor)

            _add_vba_project_bin(donor, b"nope")
            with self.assertRaises(RuntimeError):
                embed_vba_module_into_xlsm(excel_path=target, template_excel=donor, overwrite=True)


if __name__ == "__main__":
    unittest.main()

