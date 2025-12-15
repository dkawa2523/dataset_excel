import sys
import tempfile
import unittest
from contextlib import redirect_stdout
from io import StringIO
from pathlib import Path
from unittest.mock import patch

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "src"
sys.path.insert(0, str(SRC))

from clearml_dataset_excel.cli import main  # noqa: E402


class TestCliRegisterBaseExcelPreservesTemplate(unittest.TestCase):
    def test_register_base_excel_preserves_extra_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            spec_path = base / "spec.yaml"
            spec_path.write_text(
                "\n".join(
                    [
                        "schema_version: 1",
                        "addin:",
                        "  enabled: true",
                        "condition:",
                        "  columns:",
                        "    - {name: id, type: str}",
                        "    - {name: meas_path, type: path}",
                        "files:",
                        "  - id: m",
                        "    path_column: meas_path",
                        "    mapping: {axes: {t: time}, targets: [{name: f, source: value}]}",
                    ]
                )
                + "\n",
                encoding="utf-8",
            )

            (base / "clearml_dataset_excel_runner.exe").write_text("dummy", encoding="utf-8")

            import openpyxl

            base_xlsm = base / "base.xlsm"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Conditions"
            ws.append(["id", "meas_path"])
            ws.append(["old", "old.csv"])
            wb.create_sheet("Extra")["A1"].value = "keep"
            wb.save(base_xlsm)

            def _fake_upload_dataset(*, stage_dir, spec, **kwargs):  # type: ignore[no-untyped-def]
                tmpl = Path(stage_dir) / "template" / spec.template.template_filename
                self.assertTrue(tmpl.exists())
                self.assertTrue((Path(stage_dir) / "template" / "clearml_dataset_excel_runner.exe").exists())

                import json

                payload = json.loads((Path(stage_dir) / "payload.json").read_text(encoding="utf-8"))
                self.assertEqual(payload.get("runner_exe_windows"), "template/clearml_dataset_excel_runner.exe")
                wb2 = openpyxl.load_workbook(tmpl)
                self.assertIn("Extra", wb2.sheetnames)
                self.assertEqual(wb2["Conditions"].max_row, 1)
                return "ds_fake"

            with patch("clearml_dataset_excel.cli.upload_dataset", side_effect=_fake_upload_dataset):
                buf = StringIO()
                with redirect_stdout(buf):
                    rc = main(
                        [
                            "register",
                            "--spec",
                            spec_path.as_posix(),
                            "--dataset-project",
                            "P",
                            "--dataset-name",
                            "N",
                            "--base-excel",
                            base_xlsm.as_posix(),
                        ]
                    )

            self.assertEqual(rc, 0)
            self.assertIn("ds_fake", buf.getvalue())


if __name__ == "__main__":
    unittest.main()
