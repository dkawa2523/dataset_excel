import json
import sys
import tempfile
import unittest
from contextlib import redirect_stdout
from io import StringIO
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "src"
sys.path.insert(0, str(SRC))

from clearml_dataset_excel.cli import main  # noqa: E402
from clearml_dataset_excel.format_excel import generate_windows_addin_xlam  # noqa: E402
from clearml_dataset_excel import __version__ as _ADDIN_VERSION  # noqa: E402


class TestCliAddinInspect(unittest.TestCase):
    def test_addin_inspect_json(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            xlsm = root / "t.xlsm"

            import openpyxl

            wb = openpyxl.Workbook()
            wb.active.title = "Conditions"
            info = wb.create_sheet("Info")
            info["A1"].value = "dataset_project"
            info["B1"].value = "P"
            info["A2"].value = "dataset_name"
            info["B2"].value = "N"

            meta = wb.create_sheet("_meta")
            meta["A1"].value = "schema_version"
            meta["B1"].value = 1
            meta["A2"].value = "addin_enabled"
            meta["B2"].value = True
            meta["A3"].value = "addin_target_os"
            meta["B3"].value = "auto"
            meta["A4"].value = "addin_spec_filename"
            meta["B4"].value = "run.yaml"
            meta.sheet_state = "hidden"

            wb.save(xlsm)

            buf = StringIO()
            with redirect_stdout(buf):
                rc = main(["addin", "inspect", "--excel", xlsm.as_posix(), "--json"])
            self.assertEqual(rc, 0)

            payload = json.loads(buf.getvalue())
            self.assertEqual(payload["suffix"], ".xlsm")
            self.assertEqual(payload["meta_sheet"], "_meta")
            self.assertIn("Conditions", payload["sheetnames"])
            self.assertEqual(payload["meta"].get("addin_target_os"), "auto")
            self.assertEqual(payload["meta"].get("addin_spec_filename"), "run.yaml")
            self.assertIn("dataset_project", payload["info"])
            self.assertEqual(payload["info"]["dataset_project"], "P")
            self.assertFalse(payload.get("has_clearml_macro"))

    def test_addin_inspect_detects_clearml_macro(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            xlsm = root / "t.xlsm"

            import openpyxl
            import zipfile

            wb = openpyxl.Workbook()
            wb.active.title = "Conditions"
            wb.create_sheet("_meta")["A1"].value = "schema_version"
            wb.save(xlsm)

            with zipfile.ZipFile(xlsm, "a") as z:
                z.writestr("xl/vbaProject.bin", b"dummy ClearMLDatasetExcel_Run dummy")

            buf = StringIO()
            with redirect_stdout(buf):
                rc = main(["addin", "inspect", "--excel", xlsm.as_posix(), "--json"])
            self.assertEqual(rc, 0)

            payload = json.loads(buf.getvalue())
            self.assertTrue(payload.get("has_vba_project"))
            self.assertTrue(payload.get("has_clearml_macro"))

    def test_addin_inspect_xlam_reports_custom_ui_and_version(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            xlam = root / "addin.xlam"
            generate_windows_addin_xlam(xlam, overwrite=True)

            buf = StringIO()
            with redirect_stdout(buf):
                rc = main(["addin", "inspect", "--excel", xlam.as_posix(), "--json"])
            self.assertEqual(rc, 0)

            payload = json.loads(buf.getvalue())
            self.assertEqual(payload["suffix"], ".xlam")
            self.assertTrue(payload.get("has_vba_project"))
            self.assertTrue(payload.get("has_clearml_macro"))
            self.assertTrue(payload.get("has_clearml_ribbon_macro"))
            self.assertEqual(payload.get("addin_macro_version"), _ADDIN_VERSION)
            self.assertEqual(payload.get("custom_ui", {}).get("onAction"), "ClearMLDatasetExcel_Run_Ribbon")


if __name__ == "__main__":
    unittest.main()
