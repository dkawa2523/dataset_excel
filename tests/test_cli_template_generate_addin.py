import sys
import tempfile
import unittest
from contextlib import redirect_stdout
from io import StringIO
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "src"
sys.path.insert(0, str(SRC))

from clearml_dataset_excel.cli import main  # noqa: E402
from clearml_dataset_excel import __version__ as _ADDIN_VERSION  # noqa: E402
from clearml_dataset_excel.vba_project import vba_project_has_symbol  # noqa: E402


class TestCliTemplateGenerateAddin(unittest.TestCase):
    def test_template_generate_writes_vba_and_copies_spec(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            spec_dir = base / "spec"
            out_dir = base / "out"
            spec_dir.mkdir()
            out_dir.mkdir()

            spec_path = spec_dir / "my_spec.yaml"
            spec_path.write_text(
                "\n".join(
                    [
                        "schema_version: 1",
                        "addin:",
                        "  enabled: true",
                        "  target_os: mac",
                        "  spec_filename: spec_copy.yaml",
                        "  vba_module_filename: addin.bas",
                        "  command_mac: 'echo mac ${SPEC} ${EXCEL}'",
                        "  command_windows: 'echo win ${SPEC} ${EXCEL}'",
                        "condition:",
                        "  columns:",
                        "    - {name: id, type: str}",
                        "    - {name: meas_path, type: path}",
                        "files:",
                        "  - id: m",
                        "    path_column: meas_path",
                        "    mapping: {axes: {t: time}, targets: [{name: f, source: value}]}",
                    ]
                )
                + "\n",
                encoding="utf-8",
            )

            out_xlsm = out_dir / "template.xlsm"
            with redirect_stdout(StringIO()):
                rc = main(["template", "generate", "--spec", spec_path.as_posix(), "--output", out_xlsm.as_posix()])
            self.assertEqual(rc, 0)
            self.assertTrue(out_xlsm.exists())
            self.assertTrue((out_dir / "spec_copy.yaml").exists())
            self.assertTrue((out_dir / "addin.bas").exists())

            import openpyxl

            wb = openpyxl.load_workbook(out_xlsm, keep_vba=True)
            meta = wb["_meta"]
            info = wb["Info"]
            self.assertEqual(meta["A12"].value, "addin_version")
            self.assertEqual(meta["B12"].value, _ADDIN_VERSION)
            self.assertEqual(info["A17"].value, "addin_version")
            self.assertEqual(info["B17"].value, _ADDIN_VERSION)

    def test_template_generate_with_base_excel_preserves_extra_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            spec_dir = base / "spec"
            out_dir = base / "out"
            spec_dir.mkdir()
            out_dir.mkdir()

            spec_path = spec_dir / "my_spec.yaml"
            spec_path.write_text(
                "\n".join(
                    [
                        "schema_version: 1",
                        "addin:",
                        "  enabled: true",
                        "  spec_filename: spec_copy.yaml",
                        "  vba_module_filename: addin.bas",
                        "condition:",
                        "  columns:",
                        "    - {name: id, type: str}",
                        "    - {name: meas_path, type: path}",
                        "files:",
                        "  - id: m",
                        "    path_column: meas_path",
                        "    mapping: {axes: {t: time}, targets: [{name: f, source: value}]}",
                    ]
                )
                + "\n",
                encoding="utf-8",
            )

            import openpyxl

            base_xlsm = base / "base.xlsm"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Conditions"
            ws.append(["id", "meas_path"])
            ws.append(["old", "old.csv"])
            wb.create_sheet("Extra")["A1"].value = "keep"
            wb.save(base_xlsm)

            out_xlsm = out_dir / "template.xlsm"
            with redirect_stdout(StringIO()):
                rc = main(
                    [
                        "template",
                        "generate",
                        "--spec",
                        spec_path.as_posix(),
                        "--output",
                        out_xlsm.as_posix(),
                        "--base-excel",
                        base_xlsm.as_posix(),
                    ]
                )
            self.assertEqual(rc, 0)
            self.assertTrue(out_xlsm.exists())
            self.assertTrue((out_dir / "spec_copy.yaml").exists())
            self.assertTrue((out_dir / "addin.bas").exists())

            wb2 = openpyxl.load_workbook(out_xlsm)
            self.assertIn("Extra", wb2.sheetnames)
            self.assertEqual(wb2["Conditions"].max_row, 1)

    def test_template_generate_embed_vba_via_template_excel(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            spec_dir = base / "spec"
            out_dir = base / "out"
            spec_dir.mkdir()
            out_dir.mkdir()

            import openpyxl
            import zipfile

            donor_xlsm = spec_dir / "donor.xlsm"
            openpyxl.Workbook().save(donor_xlsm)
            with zipfile.ZipFile(donor_xlsm, "a") as z:
                z.writestr("xl/vbaProject.bin", b"dummy ClearMLDatasetExcel_Run dummy")

            spec_path = spec_dir / "my_spec.yaml"
            spec_path.write_text(
                "\n".join(
                    [
                        "schema_version: 1",
                        "addin:",
                        "  enabled: true",
                        "  embed_vba: true",
                        "  vba_template_excel: donor.xlsm",
                        "condition:",
                        "  columns:",
                        "    - {name: id, type: str}",
                        "    - {name: meas_path, type: path}",
                        "files:",
                        "  - id: m",
                        "    path_column: meas_path",
                        "    mapping: {axes: {t: time}, targets: [{name: f, source: value}]}",
                    ]
                )
                + "\n",
                encoding="utf-8",
            )

            out_xlsm = out_dir / "template.xlsm"
            with redirect_stdout(StringIO()):
                rc = main(["template", "generate", "--spec", spec_path.as_posix(), "--output", out_xlsm.as_posix()])
            self.assertEqual(rc, 0)

            with zipfile.ZipFile(out_xlsm, "r") as z:
                self.assertIn("xl/vbaProject.bin", set(z.namelist()))

    def test_template_generate_embed_vba_uses_bundled_default(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            spec_dir = base / "spec"
            out_dir = base / "out"
            spec_dir.mkdir()
            out_dir.mkdir()

            spec_path = spec_dir / "my_spec.yaml"
            spec_path.write_text(
                "\n".join(
                    [
                        "schema_version: 1",
                        "addin:",
                        "  enabled: true",
                        "  embed_vba: true",
                        "condition:",
                        "  columns:",
                        "    - {name: id, type: str}",
                        "    - {name: meas_path, type: path}",
                        "files:",
                        "  - id: m",
                        "    path_column: meas_path",
                        "    mapping: {axes: {t: time}, targets: [{name: f, source: value}]}",
                    ]
                )
                + "\n",
                encoding="utf-8",
            )

            out_xlsm = out_dir / "template.xlsm"
            with redirect_stdout(StringIO()):
                rc = main(["template", "generate", "--spec", spec_path.as_posix(), "--output", out_xlsm.as_posix()])
            self.assertEqual(rc, 0)

            import zipfile

            with zipfile.ZipFile(out_xlsm, "r") as z:
                vba = z.read("xl/vbaProject.bin")
            self.assertTrue(vba_project_has_symbol(vba, "ClearMLDatasetExcel_Run"))


if __name__ == "__main__":
    unittest.main()
