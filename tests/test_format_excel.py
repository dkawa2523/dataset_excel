import sys
import tempfile
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "src"
sys.path.insert(0, str(SRC))

from clearml_dataset_excel.format_excel import annotate_template_with_clearml_info, generate_condition_template  # noqa: E402
from clearml_dataset_excel.format_spec import load_format_spec  # noqa: E402


class TestFormatExcel(unittest.TestCase):
    def test_generate_condition_template(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            spec_path = base / "spec.yaml"
            spec_path.write_text(
                "\n".join(
                    [
                        "schema_version: 1",
                        "clearml:",
                        "  dataset_project: P",
                        "  dataset_name: N",
                        "template:",
                        "  condition_sheet: Conditions",
                        "  meta_sheet: _meta",
                        "condition:",
                        "  columns:",
                        "    - {name: id, type: str, required: true}",
                        "    - {name: meas_path, type: path}",
                        "files:",
                        "  - id: meas",
                        "    path_column: meas_path",
                        "    mapping: {axes: {}, targets: []}",
                    ]
                )
                + "\n",
                encoding="utf-8",
            )
            spec = load_format_spec(spec_path)
            out = base / "template.xlsm"
            generated = generate_condition_template(spec, out)
            self.assertTrue(generated.exists())

            import openpyxl

            wb = openpyxl.load_workbook(generated)
            self.assertIn("Info", wb.sheetnames)
            self.assertIn("Conditions", wb.sheetnames)
            self.assertIn("_meta", wb.sheetnames)
            ws = wb["Conditions"]
            self.assertEqual([c.value for c in ws[1]], ["id", "meas_path"])
            self.assertEqual(wb["_meta"].sheet_state, "hidden")

    def test_annotate_template_with_clearml_info(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            spec_path = base / "spec.yaml"
            spec_path.write_text(
                "\n".join(
                    [
                        "schema_version: 1",
                        "clearml:",
                        "  dataset_project: P",
                        "  dataset_name: N",
                        "condition:",
                        "  columns:",
                        "    - {name: id, type: str}",
                        "    - {name: meas_path, type: path}",
                        "files:",
                        "  - id: meas",
                        "    path_column: meas_path",
                        "    mapping: {axes: {}, targets: []}",
                    ]
                )
                + "\n",
                encoding="utf-8",
            )
            spec = load_format_spec(spec_path)
            out = base / "template.xlsm"
            generate_condition_template(spec, out)

            annotate_template_with_clearml_info(
                out,
                dataset_project="ProjX",
                dataset_name="NameY",
                dataset_id="ds123",
                clearml_web_url="https://example.com/task/123",
            )

            import openpyxl

            wb = openpyxl.load_workbook(out)
            ws = wb["Info"]
            self.assertEqual(ws["B1"].value, "ProjX")
            self.assertEqual(ws["B2"].value, "NameY")
            self.assertEqual(ws["A5"].value, "dataset_id")
            self.assertEqual(ws["B5"].value, "ds123")
            self.assertEqual(ws["B4"].value, "https://example.com/task/123")
            self.assertIsNotNone(ws["B4"].hyperlink)


if __name__ == "__main__":
    unittest.main()
