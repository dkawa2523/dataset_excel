import sys
import tempfile
import unittest
from contextlib import redirect_stderr
from io import StringIO
from pathlib import Path
from zipfile import ZipFile

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "src"
sys.path.insert(0, str(SRC))

from clearml_dataset_excel.manifest import read_rows_from_manifest  # noqa: E402
from clearml_dataset_excel.resolver import ResolvedItem, collect_local_dataset_paths, resolve_items  # noqa: E402
from clearml_dataset_excel.template import path_parts_for_template, render_dataset_path_template  # noqa: E402
from clearml_dataset_excel.wildcards import matches_any_wildcard  # noqa: E402


class TestWildcards(unittest.TestCase):
    def test_matches_any_wildcard_recursive(self) -> None:
        self.assertTrue(matches_any_wildcard("dir/file.txt", "*.txt", recursive=True))
        self.assertFalse(matches_any_wildcard("dir/file.txt", "*.jpg", recursive=True))

    def test_matches_any_wildcard_non_recursive(self) -> None:
        self.assertFalse(matches_any_wildcard("dir/file.txt", "*.txt", recursive=False))
        self.assertTrue(matches_any_wildcard("file.txt", "*.txt", recursive=False))


class TestTemplate(unittest.TestCase):
    def test_url_parts(self) -> None:
        parts = path_parts_for_template("s3://bucket/path/to/file.txt", source_path=None, base_dir=None)
        self.assertEqual(parts["basename"], "file.txt")
        self.assertEqual(parts["stem"], "file")
        self.assertEqual(parts["suffix"], ".txt")
        self.assertEqual(parts["relpath"], "path/to/file.txt")
        self.assertEqual(parts["scheme"], "s3")
        self.assertEqual(parts["netloc"], "bucket")

    def test_render_template_missing_key(self) -> None:
        with self.assertRaises(ValueError):
            render_dataset_path_template(
                "{missing}",
                {"split": "train"},
                row_index=1,
                source_text="/tmp/a.txt",
                source_path=Path("/tmp/a.txt"),
                base_dir=None,
            )


class TestResolver(unittest.TestCase):
    def test_resolve_items_skip_missing(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            rows = [{"path": str(base / "missing.txt")}]
            with redirect_stderr(StringIO()):
                items, skipped = resolve_items(
                    rows,
                    path_col="path",
                    dataset_path_col=None,
                    dataset_path_template=None,
                    base_dir=None,
                    skip_missing=True,
                )
            self.assertEqual(items, [])
            self.assertEqual(skipped, 1)

    def test_resolve_items_glob(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            base_dir = base.resolve()
            (base_dir / "data").mkdir()
            (base_dir / "data" / "a.txt").write_text("a")
            rows = [{"path": "data/*.txt", "split": "train"}]
            items, skipped = resolve_items(
                rows,
                path_col="path",
                dataset_path_col=None,
                dataset_path_template="{split}",
                base_dir=base_dir,
                skip_missing=False,
            )
            self.assertEqual(skipped, 0)
            self.assertEqual(len(items), 1)
            self.assertEqual(Path(items[0].source), base_dir / "data")
            self.assertEqual(items[0].wildcard, "*.txt")
            self.assertEqual(Path(items[0].local_base_folder), base_dir)
            self.assertEqual(items[0].dataset_path, "train")

    def test_collect_local_dataset_paths_collision(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            (base / "dir1").mkdir()
            (base / "dir2").mkdir()
            f1 = base / "dir1" / "a.txt"
            f2 = base / "dir2" / "a.txt"
            f1.write_text("1")
            f2.write_text("2")

            items = [
                ResolvedItem(source=f1.as_posix(), dataset_path=None, local_base_folder=f1.parent.as_posix(), wildcard=None),
                ResolvedItem(source=f2.as_posix(), dataset_path=None, local_base_folder=f2.parent.as_posix(), wildcard=None),
            ]
            _, collisions, matched, excluded = collect_local_dataset_paths(
                items,
                recursive=True,
                include=None,
                exclude=None,
            )
            self.assertEqual(matched, 2)
            self.assertEqual(excluded, 0)
            self.assertIn("a.txt", collisions)


class TestManifest(unittest.TestCase):
    def test_read_rows_from_manifest_csv(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            csv_path = base / "manifest.csv"
            csv_path.write_text("path,dataset_path\n/tmp/a.txt,train\n", encoding="utf-8")
            rows, cols = read_rows_from_manifest(csv_path, sheet_name=None)
            self.assertEqual(cols, ["path", "dataset_path"])
            self.assertEqual(rows[0]["dataset_path"], "train")

    def test_read_rows_from_manifest_excel_cached_formula_value(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            xlsm_path = base / "manifest.xlsm"

            import openpyxl

            wb = openpyxl.Workbook()
            ws = wb.active
            ws["A1"].value = "path"
            ws["A2"].value = "=MYADDIN()"
            wb.save(xlsm_path)

            expected = "/tmp/a.txt"
            patched = base / "manifest_cached.xlsm"
            with ZipFile(xlsm_path, "r") as zin, ZipFile(patched, "w") as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    if item.filename == "xl/worksheets/sheet1.xml":
                        xml = data.decode("utf-8")
                        xml = xml.replace(
                            '<c r="A2"><f>MYADDIN()</f><v /></c>',
                            f'<c r="A2" t="str"><f>MYADDIN()</f><v>{expected}</v></c>',
                        )
                        data = xml.encode("utf-8")
                    zout.writestr(item, data)

            rows, cols = read_rows_from_manifest(patched, sheet_name=None)
            self.assertEqual(cols, ["path"])
            self.assertEqual(rows, [{"path": expected}])


if __name__ == "__main__":
    unittest.main()
