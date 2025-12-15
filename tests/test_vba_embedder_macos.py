import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "src"
sys.path.insert(0, str(SRC))

from clearml_dataset_excel.vba_embedder import embed_vba_module_into_xlsm  # noqa: E402


class TestVbaEmbedderMacOS(unittest.TestCase):
    @unittest.skipUnless(sys.platform == "darwin", "macOS-only behavior")
    def test_embed_calls_osascript_on_macos(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)

            import openpyxl

            xlsm = base / "book.xlsm"
            wb = openpyxl.Workbook()
            wb.save(xlsm)

            bas = base / "addin.bas"
            bas.write_text('Attribute VB_Name = "ClearMLDatasetExcelAddin"\nSub ClearMLDatasetExcel_Run()\nEnd Sub\n')

            calls: list[list[str]] = []

            def _fake_run(args, **kwargs):  # type: ignore[no-untyped-def]
                calls.append([str(x) for x in args])
                class _R:
                    returncode = 0
                    stdout = ""
                    stderr = ""
                return _R()

            with patch("clearml_dataset_excel.vba_embedder.subprocess.run", side_effect=_fake_run):
                embed_vba_module_into_xlsm(excel_path=xlsm, bas_path=bas, overwrite=False)

            osascript_calls = [c for c in calls if c and c[0] == "osascript"]
            self.assertEqual(len(osascript_calls), 1)
            self.assertEqual(Path(osascript_calls[0][-2]).resolve(), xlsm.resolve())
            self.assertEqual(Path(osascript_calls[0][-1]).resolve(), bas.resolve())

    @unittest.skipUnless(sys.platform == "darwin", "macOS-only behavior")
    def test_embed_accessibility_error_message(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)

            import openpyxl

            xlsm = base / "book.xlsm"
            wb = openpyxl.Workbook()
            wb.save(xlsm)

            bas = base / "addin.bas"
            bas.write_text('Attribute VB_Name = "ClearMLDatasetExcelAddin"\nSub ClearMLDatasetExcel_Run()\nEnd Sub\n')

            import subprocess

            def _fail(*args, **kwargs):  # type: ignore[no-untyped-def]
                raise subprocess.CalledProcessError(
                    returncode=1, cmd=args[0], output="", stderr="osascriptには補助アクセスは許可されません。 (-1719)"
                )

            with patch("clearml_dataset_excel.vba_embedder.subprocess.run", side_effect=_fail):
                with self.assertRaises(RuntimeError) as ctx:
                    embed_vba_module_into_xlsm(excel_path=xlsm, bas_path=bas, overwrite=False)
            self.assertIn("Accessibility", str(ctx.exception))

    @unittest.skipUnless(sys.platform == "darwin", "macOS-only behavior")
    def test_embed_timeout_error_message(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)

            import openpyxl

            xlsm = base / "book.xlsm"
            wb = openpyxl.Workbook()
            wb.save(xlsm)

            bas = base / "addin.bas"
            bas.write_text('Attribute VB_Name = "ClearMLDatasetExcelAddin"\nSub ClearMLDatasetExcel_Run()\nEnd Sub\n')

            import subprocess

            def _fail(*args, **kwargs):  # type: ignore[no-untyped-def]
                raise subprocess.CalledProcessError(
                    returncode=1,
                    cmd=args[0],
                    output="",
                    stderr="Microsoft Excelでエラーが起きました: AppleEventがタイムアウトしました。 (-1712)",
                )

            with patch("clearml_dataset_excel.vba_embedder.subprocess.run", side_effect=_fail):
                with self.assertRaises(RuntimeError) as ctx:
                    embed_vba_module_into_xlsm(excel_path=xlsm, bas_path=bas, overwrite=False)
            self.assertIn("timed out", str(ctx.exception).lower())

    @unittest.skipUnless(sys.platform == "darwin", "macOS-only behavior")
    def test_embed_skips_when_symbol_exists_and_no_overwrite(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            xlsm = base / "book.xlsm"

            import zipfile

            with zipfile.ZipFile(xlsm, "w") as z:
                z.writestr("xl/vbaProject.bin", b"ClearMLDatasetExcel_Run")

            bas = base / "addin.bas"
            bas.write_text('Attribute VB_Name = "ClearMLDatasetExcelAddin"\nSub ClearMLDatasetExcel_Run()\nEnd Sub\n')

            with patch("clearml_dataset_excel.vba_embedder.subprocess.run") as run:
                embed_vba_module_into_xlsm(excel_path=xlsm, bas_path=bas, overwrite=False)
            run.assert_not_called()

    @unittest.skipUnless(sys.platform == "darwin", "macOS-only behavior")
    def test_embed_raises_when_symbol_exists_and_overwrite(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            xlsm = base / "book.xlsm"

            import zipfile

            with zipfile.ZipFile(xlsm, "w") as z:
                z.writestr("xl/vbaProject.bin", b"ClearMLDatasetExcel_Run")

            bas = base / "addin.bas"
            bas.write_text('Attribute VB_Name = "ClearMLDatasetExcelAddin"\nSub ClearMLDatasetExcel_Run()\nEnd Sub\n')

            with patch("clearml_dataset_excel.vba_embedder.subprocess.run") as run:
                with self.assertRaises(RuntimeError):
                    embed_vba_module_into_xlsm(excel_path=xlsm, bas_path=bas, overwrite=True)
            run.assert_not_called()


if __name__ == "__main__":
    unittest.main()
