import sys
import tempfile
import unittest
from contextlib import redirect_stdout
from io import StringIO
from pathlib import Path
from unittest.mock import patch

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "src"
sys.path.insert(0, str(SRC))

from clearml_dataset_excel.cli import main  # noqa: E402


class _FakeTaskNoConfig:
    def __init__(self) -> None:
        self.id = "ds_base_1"

    def get_configuration_object_as_dict(self, name: str):  # type: ignore[no-untyped-def]
        return {}


class _FakeBaseDataset:
    def __init__(self, local_copy: Path) -> None:
        self._local_copy = local_copy

    def get_local_copy(self, *args, **kwargs):  # type: ignore[no-untyped-def]
        return self._local_copy.as_posix()


class TestCliAgentReprocessNoUploadStagePersists(unittest.TestCase):
    def test_agent_reprocess_no_upload_keeps_stage_dir(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            (root / "spec").mkdir()
            input_dir = root / "input"
            (input_dir / "external").mkdir(parents=True)
            template_dir = root / "template"
            template_dir.mkdir()

            # Dataset includes spec yaml
            spec_path = root / "spec" / "spec.yaml"
            spec_path.write_text(
                "\n".join(
                    [
                        "schema_version: 1",
                        "condition:",
                        "  columns:",
                        "    - {name: id, type: str}",
                        "    - {name: meas_path, type: path}",
                        "files:",
                        "  - id: m",
                        "    path_column: meas_path",
                        "    format: csv",
                        "    mapping: {axes: {t: time}, targets: [{name: f, source: value}]}",
                    ]
                )
                + "\n",
                encoding="utf-8",
            )

            meas = input_dir / "external" / "000_meas.csv"
            meas.write_text("time,value\n0,1\n1,2\n", encoding="utf-8")

            import openpyxl

            xlsm = input_dir / "conditions.xlsm"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Conditions"
            ws.append(["id", "meas_path"])
            ws.append(["s1", "/abs/meas.csv"])
            wb.save(xlsm)

            # Base dataset includes template Excel (to preserve VBA/extra sheets)
            tmpl = template_dir / "condition_template.xlsm"
            wb_t = openpyxl.Workbook()
            ws_t = wb_t.active
            ws_t.title = "Conditions"
            ws_t.append(["id", "meas_path"])
            wb_t.create_sheet("Extra")["A1"].value = "keep"
            wb_t.save(tmpl)

            (root / "payload.json").write_text(
                '{"payload_version":1,"spec_path":"spec/spec.yaml","template_excel":"template/condition_template.xlsm","condition_excel":"input/conditions.xlsm","path_map":{"/abs/meas.csv":"input/external/000_meas.csv"}}\n',
                encoding="utf-8",
            )

            out_root = root / "out"
            out_root.mkdir()

            fake_task = _FakeTaskNoConfig()
            base_ds = _FakeBaseDataset(root)

            def _dataset_get_side_effect(**kwargs):  # type: ignore[no-untyped-def]
                return base_ds

            with patch("clearml.Task") as task_cls, patch("clearml.Dataset") as dataset_cls:
                task_cls.current_task.return_value = fake_task
                dataset_cls.get.side_effect = _dataset_get_side_effect

                buf = StringIO()
                with redirect_stdout(buf):
                    rc = main(["agent", "reprocess", "--no-upload", "--output-root", out_root.as_posix()])

            self.assertEqual(rc, 0)
            lines = [ln.strip() for ln in buf.getvalue().splitlines() if ln.strip()]
            self.assertGreaterEqual(len(lines), 2)
            output_dir = Path(lines[0]).expanduser().resolve()
            stage_dir = Path(lines[1]).expanduser().resolve()
            self.assertTrue(output_dir.exists())
            self.assertTrue(stage_dir.exists())
            self.assertTrue((stage_dir / "payload.json").exists())
            self.assertTrue(str(stage_dir).startswith(str(output_dir)))

            import openpyxl

            staged_template = stage_dir / "template" / "condition_template.xlsm"
            self.assertTrue(staged_template.exists())
            wb2 = openpyxl.load_workbook(staged_template)
            self.assertIn("Extra", wb2.sheetnames)


if __name__ == "__main__":
    unittest.main()
