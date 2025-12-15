from __future__ import annotations

import shutil
from datetime import datetime, timezone
from pathlib import Path

from .format_spec import DatasetFormatSpec
from . import __version__ as _ADDIN_VERSION
from .utils import clear_macos_quarantine


def _repair_vba_metadata_if_present(excel_path: Path) -> None:
    """
    Best-effort: repair OOXML metadata required for VBA projects.

    Some writers (notably openpyxl keep_vba=True) can drop the vbaProject override from
    [Content_Types].xml, which may cause Excel to show an empty macro list.
    """
    try:
        import zipfile

        with zipfile.ZipFile(excel_path, "r") as z:
            if "xl/vbaProject.bin" not in set(z.namelist()):
                return
    except Exception:
        return

    try:
        from .vba_embedder import embed_vba_module_into_xlsm

        embed_vba_module_into_xlsm(excel_path=excel_path, bas_path=None, overwrite=False, template_excel=None)
    except Exception:
        # Best-effort repair; do not fail template generation.
        return


def _excel_number_format(dtype: str) -> str | None:
    dt = dtype.lower().strip()
    if dt in {"int", "int64", "integer"}:
        return "0"
    if dt in {"float", "float64", "number"}:
        return "0.########"
    if dt in {"bool", "boolean"}:
        return "@"
    if dt in {"date"}:
        return "yyyy-mm-dd"
    if dt in {"datetime"}:
        return "yyyy-mm-dd hh:mm:ss"
    return None


def _populate_template_workbook(
    wb,  # type: ignore[no-untyped-def]
    spec: DatasetFormatSpec,
    *,
    clear_conditions_data: bool,
) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    info_ws = wb["Info"] if "Info" in wb.sheetnames else wb.create_sheet("Info")
    cond_ws = (
        wb[spec.template.condition_sheet]
        if spec.template.condition_sheet in wb.sheetnames
        else wb.create_sheet(spec.template.condition_sheet)
    )
    meta_ws = (
        wb[spec.template.meta_sheet]
        if spec.template.meta_sheet in wb.sheetnames
        else wb.create_sheet(spec.template.meta_sheet)
    )

    # Info sheet (visible)
    info_ws["A1"].value = "dataset_project"
    info_ws["A2"].value = "dataset_name"
    info_ws["A3"].value = "template_generated_at"
    info_ws["A4"].value = "clearml_web_url"
    info_ws["A5"].value = "dataset_id"
    info_ws["A6"].value = "How to use"
    info_ws["A7"].value = "1) Fill rows in the Conditions sheet"
    info_ws["A8"].value = "2) Run: clearml-dataset-excel run --spec <spec.yaml> --excel <filled.xlsx/.xlsm>"
    info_ws["A10"].value = "addin_enabled"
    info_ws["A11"].value = "addin_target_os"
    info_ws["A12"].value = "addin_spec_filename"
    info_ws["A13"].value = "vba_module_filename"
    info_ws["A14"].value = "addin_windows_mode"
    info_ws["A15"].value = "addin_windows_template_filename"
    info_ws["A16"].value = "addin_windows_addin_filename"
    info_ws["A17"].value = "addin_version"

    if spec.clearml:
        info_ws["B1"].value = spec.clearml.dataset_project
        info_ws["B2"].value = spec.clearml.dataset_name
    info_ws["B3"].value = datetime.now(timezone.utc).isoformat()
    if not info_ws["B4"].value:
        info_ws["B4"].value = ""
    if not info_ws["B5"].value:
        info_ws["B5"].value = ""
    info_ws["B10"].value = bool(spec.addin.enabled)
    info_ws["B11"].value = spec.addin.target_os
    info_ws["B12"].value = spec.addin.spec_filename or ""
    info_ws["B13"].value = spec.addin.vba_module_filename
    info_ws["B14"].value = spec.addin.windows_mode
    info_ws["B15"].value = spec.addin.windows_template_filename or ""
    info_ws["B16"].value = spec.addin.windows_addin_filename
    info_ws["B17"].value = _ADDIN_VERSION

    # Meta sheet (hidden)
    meta_ws["A1"].value = "schema_version"
    meta_ws["B1"].value = spec.schema_version
    meta_ws["A2"].value = "condition_sheet"
    meta_ws["B2"].value = spec.template.condition_sheet
    meta_ws["A3"].value = "addin_enabled"
    meta_ws["B3"].value = bool(spec.addin.enabled)
    meta_ws["A4"].value = "addin_target_os"
    meta_ws["B4"].value = spec.addin.target_os
    meta_ws["A5"].value = "addin_spec_filename"
    meta_ws["B5"].value = spec.addin.spec_filename or ""
    meta_ws["A6"].value = "addin_command"
    meta_ws["B6"].value = spec.addin.command or ""
    meta_ws["A7"].value = "addin_command_mac"
    meta_ws["B7"].value = spec.addin.command_mac or ""
    meta_ws["A8"].value = "addin_command_windows"
    meta_ws["B8"].value = spec.addin.command_windows or ""
    meta_ws["A9"].value = "addin_windows_mode"
    meta_ws["B9"].value = spec.addin.windows_mode
    meta_ws["A10"].value = "addin_windows_template_filename"
    meta_ws["B10"].value = spec.addin.windows_template_filename or ""
    meta_ws["A11"].value = "addin_windows_addin_filename"
    meta_ws["B11"].value = spec.addin.windows_addin_filename
    meta_ws["A12"].value = "addin_version"
    meta_ws["B12"].value = _ADDIN_VERSION

    meta_ws.sheet_state = "hidden"

    # Conditions header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    required_fill = PatternFill("solid", fgColor="C00000")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cond_ws.freeze_panes = "A2"
    if clear_conditions_data and cond_ws.max_row > 1:
        cond_ws.delete_rows(2, cond_ws.max_row - 1)

    names = [c.name for c in spec.condition_columns]
    max_existing = int(cond_ws.max_column or 0)
    max_cols = max(max_existing, len(names))
    for col_idx in range(1, max_cols + 1):
        cell = cond_ws.cell(row=1, column=col_idx)
        if col_idx <= len(names):
            cell.value = names[col_idx - 1]
        else:
            cell.value = None

    for col_idx, col in enumerate(spec.condition_columns, start=1):
        cell = cond_ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = required_fill if col.required else header_fill
        cell.alignment = header_alignment

        # Set a reasonable width
        width = max(12, min(40, len(col.name) + 2))
        if col.dtype.lower().strip() in {"path", "str", "string"} or col.name.lower().endswith("_path"):
            width = max(width, 28)
        cond_ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # Optional: add data validation for enum/bool
    try:
        from openpyxl.worksheet.datavalidation import DataValidation

        for col_idx, col in enumerate(spec.condition_columns, start=1):
            dt = col.dtype.lower().strip()
            if dt in {"bool", "boolean"}:
                dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=not col.required)
                cond_ws.add_data_validation(dv)
                dv.add(f"{openpyxl.utils.get_column_letter(col_idx)}2:{openpyxl.utils.get_column_letter(col_idx)}1048576")
            elif col.enum:
                # Excel list validation has length limits; keep it simple for now.
                items = ",".join([x.replace(",", " ") for x in col.enum])
                if 1 <= len(items) <= 200:
                    dv = DataValidation(type="list", formula1=f'"{items}"', allow_blank=not col.required)
                    cond_ws.add_data_validation(dv)
                    dv.add(
                        f"{openpyxl.utils.get_column_letter(col_idx)}2:{openpyxl.utils.get_column_letter(col_idx)}1048576"
                    )
    except Exception:
        # Data validation is optional; do not fail template generation.
        pass


def generate_condition_template_from_excel(
    base_excel: str | Path,
    spec: DatasetFormatSpec,
    output_path: str | Path,
    *,
    overwrite: bool = False,
    clear_conditions_data: bool = True,
) -> Path:
    """
    Generate a condition Excel template by copying an existing workbook (preserves VBA if present).
    """
    base = Path(base_excel).expanduser().resolve()
    out = Path(output_path).expanduser().resolve()
    if not base.exists():
        raise FileNotFoundError(f"Base Excel not found: {base}")
    if out.exists() and not overwrite:
        raise FileExistsError(f"Output already exists: {out}")

    try:
        import openpyxl
    except Exception as e:  # pragma: no cover
        raise RuntimeError("Generating Excel requires 'openpyxl'. Install dependencies first.") from e

    out.parent.mkdir(parents=True, exist_ok=True)
    if base.resolve() != out.resolve():
        shutil.copy2(base, out)

    wb = openpyxl.load_workbook(out, keep_vba=True)
    _populate_template_workbook(wb, spec, clear_conditions_data=clear_conditions_data)
    wb.save(out)
    _repair_vba_metadata_if_present(out)
    clear_macos_quarantine(out)
    return out


def annotate_template_with_clearml_info(
    template_path: str | Path,
    *,
    dataset_project: str | None,
    dataset_name: str | None,
    dataset_id: str | None,
    clearml_web_url: str | None,
) -> None:
    """
    Best-effort: write ClearML identifiers/links into the visible Info sheet.
    Keeps VBA (if any) by loading with keep_vba=True.
    """
    path = Path(template_path).expanduser().resolve()
    if not path.exists():
        return

    try:
        import openpyxl
    except Exception:  # pragma: no cover
        return

    wb = openpyxl.load_workbook(path, keep_vba=True)
    try:
        if "Info" not in wb.sheetnames:
            return
        ws = wb["Info"]

        if dataset_project is not None:
            ws["A1"].value = "dataset_project"
            ws["B1"].value = str(dataset_project)
        if dataset_name is not None:
            ws["A2"].value = "dataset_name"
            ws["B2"].value = str(dataset_name)

        ws["A5"].value = "dataset_id"
        ws["B5"].value = "" if dataset_id is None else str(dataset_id)

        ws["A4"].value = "clearml_web_url"
        if clearml_web_url:
            ws["B4"].value = str(clearml_web_url)
            try:
                ws["B4"].hyperlink = str(clearml_web_url)
                ws["B4"].style = "Hyperlink"
            except Exception:
                pass

        wb.save(path)
    finally:
        try:
            vba_archive = getattr(wb, "vba_archive", None)
            if vba_archive is not None:
                vba_archive.close()
        except Exception:
            pass
    _repair_vba_metadata_if_present(path)
    clear_macos_quarantine(path)


def generate_condition_template(spec: DatasetFormatSpec, output_path: str | Path, *, overwrite: bool = False) -> Path:
    """
    Generate a condition Excel template from DatasetFormatSpec.

    Note: This generates a macro-enabled file extension (.xlsm) but does not embed VBA automatically.
    """
    out = Path(output_path).expanduser().resolve()
    if out.exists() and not overwrite:
        raise FileExistsError(f"Output already exists: {out}")

    try:
        import openpyxl
    except Exception as e:  # pragma: no cover
        raise RuntimeError("Generating Excel requires 'openpyxl'. Install dependencies first.") from e

    wb = openpyxl.Workbook()
    # Remove default sheet
    default_ws = wb.active
    wb.remove(default_ws)
    _populate_template_workbook(wb, spec, clear_conditions_data=False)

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    clear_macos_quarantine(out)
    return out


def generate_windows_addin_xlam(output_path: str | Path, *, overwrite: bool = False) -> Path:
    """
    Generate a Windows Excel add-in file (.xlam) that contains ClearMLDatasetExcel_Run.

    Note: This uses the bundled default vbaProject.bin (no Excel/COM/UI automation required).
    """
    out = Path(output_path).expanduser().resolve()
    if out.suffix.lower() != ".xlam":
        raise ValueError(f"Output must be .xlam: {out}")
    if out.exists() and not overwrite:
        raise FileExistsError(f"Output already exists: {out}")

    try:
        import openpyxl
    except Exception as e:  # pragma: no cover
        raise RuntimeError("Generating .xlam requires 'openpyxl'. Install dependencies first.") from e

    out.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    try:
        ws = wb.active
        ws.title = "Addin"
    except Exception:
        pass
    wb.save(out)

    from .vba_embedder import embed_vba_module_into_xlsm

    embed_vba_module_into_xlsm(
        excel_path=out,
        bas_path=None,
        overwrite=True,
        template_excel=None,
    )
    clear_macos_quarantine(out)
    return out
