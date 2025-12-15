from __future__ import annotations

from pathlib import Path
from typing import Any

import re

from .vba_project import vba_project_has_symbol
from .msovba import decompress_stream
from .utils import get_macos_quarantine


_CUSTOM_UI_PART = "customUI/customUI.xml"


def _extract_vba_module_source(vba_project_bin: bytes, module_name: str) -> str | None:
    try:
        import io

        import olefile  # type: ignore[import-not-found]

        ole = olefile.OleFileIO(io.BytesIO(vba_project_bin))
    except Exception:
        return None

    try:
        try:
            stream = ole.openstream(["VBA", module_name]).read()
        except Exception:
            return None

        idx = stream.lower().find(b"\x00attribut")
        if idx < 0 or idx < 3:
            return None
        start = idx - 3
        if start < 0 or stream[start] != 0x01:
            return None

        src = decompress_stream(stream[start:])
        return src.decode("cp1252", errors="replace")
    finally:
        try:
            ole.close()
        except Exception:
            pass


def _extract_addin_macro_version(vba_project_bin: bytes) -> str | None:
    # Generated .bas uses ADDIN_VERSION; bundled default uses V$ in Calculations.
    for module_name in ("ClearMLDatasetExcelAddin", "Calculations"):
        src = _extract_vba_module_source(vba_project_bin, module_name)
        if not src:
            continue

        m = re.search(r'Private\s+Const\s+ADDIN_VERSION\s+As\s+String\s*=\s*"([^"]+)"', src)
        if m:
            return m.group(1)

        m = re.search(r'Private\s+Const\s+V\$?\s*=\s*"([^"]+)"', src)
        if m:
            return m.group(1)
    return None


def inspect_addin_excel(
    excel_path: str | Path,
    *,
    meta_sheet_name: str = "_meta",
    info_sheet_name: str = "Info",
) -> dict[str, Any]:
    """
    Inspect a condition template Excel file for add-in usage.

    This is a best-effort helper for validating that:
    - the workbook contains expected sheets
    - the hidden meta sheet contains addin_* keys
    - the file contains a VBA project (vbaProject.bin) when macros are embedded
    """
    path = Path(excel_path).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"Excel not found: {path}")

    has_vba_project = False
    has_clearml_macro = False
    has_clearml_ribbon_macro = False
    addin_macro_version: str | None = None
    vba_project_size = 0
    vba_moduleprivate_count = 0
    is_zip = False
    zip_entries: dict[str, dict[str, Any]] = {}
    custom_ui: dict[str, str] = {}
    app_properties: dict[str, str] = {}
    workbook_has_file_version = False
    try:
        import zipfile

        with zipfile.ZipFile(path) as zf:
            is_zip = True
            names = set(zf.namelist())
            has_vba_project = "xl/vbaProject.bin" in names
            if _CUSTOM_UI_PART in names:
                custom_ui["part"] = _CUSTOM_UI_PART
                try:
                    from xml.etree import ElementTree as ET

                    root = ET.fromstring(zf.read(_CUSTOM_UI_PART))
                    for elem in root.iter():
                        if not str(elem.tag).endswith("button"):
                            continue
                        if elem.attrib.get("onAction"):
                            custom_ui["onAction"] = str(elem.attrib.get("onAction"))
                        if elem.attrib.get("label"):
                            custom_ui["label"] = str(elem.attrib.get("label"))
                        if elem.attrib.get("id"):
                            custom_ui["id"] = str(elem.attrib.get("id"))
                        break
                except Exception:
                    pass
            for n in sorted(names):
                if n in {
                    "[Content_Types].xml",
                    "_rels/.rels",
                    "xl/workbook.xml",
                    "xl/_rels/workbook.xml.rels",
                    "xl/vbaProject.bin",
                } or (n.startswith("xl/worksheets/") and n.endswith(".xml")):
                    try:
                        info = zf.getinfo(n)
                        zip_entries[n] = {
                            "compress_type": int(getattr(info, "compress_type", 0) or 0),
                            "file_size": int(getattr(info, "file_size", 0) or 0),
                        }
                    except Exception:
                        zip_entries[n] = {}
            if has_vba_project:
                try:
                    data = zf.read("xl/vbaProject.bin")
                    vba_project_size = int(len(data))
                    has_clearml_macro = vba_project_has_symbol(data, "ClearMLDatasetExcel_Run")
                    has_clearml_ribbon_macro = vba_project_has_symbol(data, "ClearMLDatasetExcel_Run_Ribbon")
                    addin_macro_version = _extract_addin_macro_version(data)
                    # Extra diagnostic: MODULEPRIVATE records in VBA/dir often cause Excel's macro list to be empty.
                    try:
                        import io

                        import olefile  # type: ignore[import-not-found]

                        ole = olefile.OleFileIO(io.BytesIO(data))
                        try:
                            if ole.exists("VBA/dir"):
                                dir_comp = ole.openstream("VBA/dir").read()
                                dir_raw = decompress_stream(dir_comp)
                                pat = b"\x28\x00\x00\x00\x00\x00\x2B\x00\x00\x00\x00\x00"
                                vba_moduleprivate_count = int(dir_raw.count(pat))
                        finally:
                            ole.close()
                    except Exception:
                        vba_moduleprivate_count = 0
                except Exception:
                    has_clearml_macro = False
                    vba_project_size = 0
                    vba_moduleprivate_count = 0

            # Read small OOXML metadata that sometimes affects Excel behavior.
            try:
                if "docProps/app.xml" in names:
                    from xml.etree import ElementTree as ET

                    root = ET.fromstring(zf.read("docProps/app.xml"))
                    ns = {"a": "http://schemas.openxmlformats.org/officeDocument/2006/extended-properties"}
                    app = root.find("a:Application", ns)
                    ver = root.find("a:AppVersion", ns)
                    if app is not None and (app.text or "").strip():
                        app_properties["Application"] = (app.text or "").strip()
                    if ver is not None and (ver.text or "").strip():
                        app_properties["AppVersion"] = (ver.text or "").strip()
            except Exception:
                app_properties = {}

            try:
                if "xl/workbook.xml" in names:
                    from xml.etree import ElementTree as ET

                    root = ET.fromstring(zf.read("xl/workbook.xml"))
                    ns = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
                    workbook_has_file_version = root.find("s:fileVersion", ns) is not None
            except Exception:
                workbook_has_file_version = False
    except Exception:
        is_zip = False
        has_vba_project = False
        has_clearml_macro = False
        vba_project_size = 0

    try:
        import openpyxl
    except Exception as e:  # pragma: no cover
        raise RuntimeError("Inspecting Excel requires 'openpyxl'. Install dependencies first.") from e

    base_result = {
        "excel": path.as_posix(),
        "suffix": path.suffix.lower(),
        "is_zip": bool(is_zip),
        "has_vba_project": bool(has_vba_project),
        "has_clearml_macro": bool(has_clearml_macro),
        "has_clearml_ribbon_macro": bool(has_clearml_ribbon_macro),
        "addin_macro_version": addin_macro_version,
        "vba_project_size": int(vba_project_size),
        "vba_moduleprivate_count": int(vba_moduleprivate_count),
        "custom_ui": custom_ui,
        "macos_quarantine": get_macos_quarantine(path),
        "app_properties": app_properties,
        "workbook_has_fileVersion": bool(workbook_has_file_version),
        "zip_entries": zip_entries,
    }

    try:
        wb = openpyxl.load_workbook(path, keep_vba=True, data_only=True)
    except Exception:
        # Some file types (notably .xlam) may not be fully supported by openpyxl.
        # Still return zip-level diagnostics.
        return {
            **base_result,
            "sheetnames": [],
            "worksheet_code_names": {},
            "meta_sheet": None,
            "meta": {},
            "info": {},
        }
    try:
        sheetnames = list(wb.sheetnames)
        worksheet_code_names: dict[str, str] = {}
        for name in sheetnames:
            try:
                ws = wb[name]
                sp = getattr(ws, "sheet_properties", None)
                code = getattr(sp, "codeName", None) if sp is not None else None
                if isinstance(code, str) and code.strip():
                    worksheet_code_names[name] = code.strip()
                else:
                    worksheet_code_names[name] = ""
            except Exception:
                worksheet_code_names[name] = ""

        meta_ws = None
        meta_ws_name = None
        if meta_sheet_name in wb.sheetnames:
            meta_ws = wb[meta_sheet_name]
            meta_ws_name = meta_sheet_name
        elif "_meta" in wb.sheetnames:
            meta_ws = wb["_meta"]
            meta_ws_name = "_meta"
        else:
            # Best-effort search: find a sheet that has "schema_version" in column A.
            for name in wb.sheetnames:
                ws = wb[name]
                try:
                    v = ws["A1"].value
                    if isinstance(v, str) and v.strip() == "schema_version":
                        meta_ws = ws
                        meta_ws_name = name
                        break
                except Exception:
                    continue

        meta: dict[str, str] = {}
        if meta_ws is not None:
            empty_streak = 0
            for i in range(1, 2001):
                k = meta_ws.cell(row=i, column=1).value
                if k is None or (isinstance(k, str) and not k.strip()):
                    empty_streak += 1
                    if empty_streak >= 100:
                        break
                    continue
                empty_streak = 0
                key = str(k).strip()
                v = meta_ws.cell(row=i, column=2).value
                meta[key] = "" if v is None else str(v)

        info: dict[str, str] = {}
        if info_sheet_name in wb.sheetnames:
            ws = wb[info_sheet_name]
            # Read A1..A20 as key labels, B as values
            for i in range(1, 21):
                k = ws.cell(row=i, column=1).value
                if k is None or (isinstance(k, str) and not k.strip()):
                    continue
                key = str(k).strip()
                v = ws.cell(row=i, column=2).value
                info[key] = "" if v is None else str(v)

        return {
            **base_result,
            "sheetnames": sheetnames,
            "worksheet_code_names": worksheet_code_names,
            "meta_sheet": meta_ws_name,
            "meta": meta,
            "info": info,
        }
    finally:
        try:
            vba_archive = getattr(wb, "vba_archive", None)
            if vba_archive is not None:
                vba_archive.close()
        except Exception:
            pass
