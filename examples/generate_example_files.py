from __future__ import annotations

import shutil
import sys
from pathlib import Path


def main() -> int:
    root = Path(__file__).resolve().parent
    repo_root = root.parent
    src = repo_root / "src"
    if src.exists():
        sys.path.insert(0, src.as_posix())

    data_dir = root / "data"
    data_dir.mkdir(parents=True, exist_ok=True)

    try:
        import openpyxl
    except Exception as e:
        raise RuntimeError("Generating .xlsx requires 'openpyxl'. Install dependencies first.") from e

    # Measurement CSVs (time-series example)
    (data_dir / "meas_a_s1.csv").write_text("time,value\n0,1\n1,3\n2,5\n", encoding="utf-8")
    (data_dir / "meas_b_s1.csv").write_text("time,value\n0,10\n1,30\n2,50\n", encoding="utf-8")
    (data_dir / "meas_a_s2.csv").write_text("time,value\n0,2\n1,4\n2,6\n", encoding="utf-8")
    (data_dir / "meas_b_s2.csv").write_text("time,value\n0,20\n1,40\n2,60\n", encoding="utf-8")

    from clearml_dataset_excel.format_excel import generate_condition_template, generate_windows_addin_xlam  # noqa: E402
    from clearml_dataset_excel.format_spec import load_format_spec  # noqa: E402
    from clearml_dataset_excel.utils import clear_macos_quarantine  # noqa: E402
    from clearml_dataset_excel.vba_addin import write_vba_module  # noqa: E402
    from clearml_dataset_excel.vba_embedder import embed_vba_module_into_xlsm  # noqa: E402

    spec_path = root / "run.yaml"
    spec = load_format_spec(spec_path)

    # Generate template + .bas (add-in module)
    template_path = root / spec.template.template_filename
    generate_condition_template(spec, template_path, overwrite=True)

    if spec.addin.enabled:
        write_vba_module(
            root / spec.addin.vba_module_filename,
            meta_sheet_name=spec.template.meta_sheet,
        )
        if spec.addin.embed_vba:
            vba_template_excel = spec.addin.vba_template_excel
            vba_template_path = None
            if isinstance(vba_template_excel, str) and vba_template_excel.strip():
                cand = Path(vba_template_excel).expanduser()
                vba_template_path = cand.resolve() if cand.is_absolute() else (spec_path.parent / cand).resolve()
            embed_vba_module_into_xlsm(
                excel_path=template_path,
                overwrite=True,
                template_excel=vba_template_path,
                bas_path=None,
            )

    win_template_path = None
    win_addin_path = None
    if spec.addin.enabled and spec.addin.windows_mode == "addin":
        win_template_path = root / (spec.addin.windows_template_filename or "condition_template.xlsx")
        generate_condition_template(spec, win_template_path, overwrite=True)
        win_addin_path = root / spec.addin.windows_addin_filename
        generate_windows_addin_xlam(win_addin_path, overwrite=True)

    # Filled condition Excel based on the template (keeps Info/_meta sheets)
    out_xlsm = root / "conditions_filled.xlsm"
    shutil.copy2(template_path, out_xlsm)
    wb = openpyxl.load_workbook(out_xlsm, keep_vba=True)
    ws = wb[spec.template.condition_sheet]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    ws.append(["s1", 25.0, "data/meas_a_s1.csv", "data/meas_b_s1.csv"])
    ws.append(["s2", 30.0, "data/meas_a_s2.csv", "data/meas_b_s2.csv"])
    wb.save(out_xlsm)
    clear_macos_quarantine(out_xlsm)
    if spec.addin.enabled and spec.addin.embed_vba:
        # openpyxl keep_vba=True can drop the vbaProject override; repair the OOXML metadata after saving.
        embed_vba_module_into_xlsm(
            excel_path=out_xlsm,
            overwrite=False,
            template_excel=None,
            bas_path=None,
        )

    out_xlsx = None
    if win_template_path is not None and win_template_path.exists():
        out_xlsx = root / "conditions_filled.xlsx"
        shutil.copy2(win_template_path, out_xlsx)
        wb_x = openpyxl.load_workbook(out_xlsx, keep_vba=False)
        ws_x = wb_x[spec.template.condition_sheet]
        if ws_x.max_row > 1:
            ws_x.delete_rows(2, ws_x.max_row - 1)
        ws_x.append(["s1", 25.0, "data/meas_a_s1.csv", "data/meas_b_s1.csv"])
        ws_x.append(["s2", 30.0, "data/meas_a_s2.csv", "data/meas_b_s2.csv"])
        wb_x.save(out_xlsx)

    print(template_path.as_posix())
    if spec.addin.enabled:
        print((root / spec.addin.vba_module_filename).as_posix())
    if win_template_path is not None:
        print(win_template_path.as_posix())
    if win_addin_path is not None:
        print(win_addin_path.as_posix())
    print(out_xlsm.as_posix())
    if out_xlsx is not None:
        print(out_xlsx.as_posix())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
